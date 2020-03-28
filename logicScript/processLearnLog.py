#!/usr/bin/env python3

# requires openpyxl; install with the following:
# pip3 install openpyxl

from datetime import datetime
from datetime import date
import csv
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
import re

numSheets = 10


# Used to compute how early a student accessed a ressoruce (on the time scale from release to exam)
examDate = date(2019, 1, 18) # for excel, yyyy;mm;dd

# Used to determine time of release (needs rework for more than one lecturer)
lecturer = 'D D Freydenberger'

dateformat = "%d/%m/%y, %H:%M"
outputDateFormat = "%d/%m/%y"

#Learn logfile for module
logfile = 'log.csv'
#csv with student data (from cotutor)
regfile = 'cotutor_students_download.csv'

otherInteresting = []

def generalizedSheet(str):
    if isSheet(str):
        return str + ' with spoilers (or not)'
    if isSpoiler(str):
        return str + ' (or not)'


def isSpoiler(str):
    return (('Sheet' in str) and ('with spoilers' in str) and ('(or not)' not in str))

def isGenSheet(str):
    return (('Sheet' in str) and ('with spoilers' in str) and ('(or not)' in str))

def isSheet(str):
    return (('Sheet' in str) and ('with spoilers' not in str) and not ('Quiz' in str))

def isOther(str):
    return (('File: Lecture notes' in str) or ('Quiz:' in str) or ('File: Example exam' in str) or ('Folder: Slides' in str))

def isInteresting(str):
    return (isSpoiler(str) or isSheet(str) or isOther(str))

students = dict()
firstAccFor = dict()

def importRegsFromCoTutor(regfile):
    regstuds = dict()
    with open(regfile, newline='') as studfile:
        reader = csv.reader(studfile, delimiter=',', quotechar='"')
        for row in reader:
            regno = row[2].rstrip(' ')
            fn = row[1].rstrip(' ')
            sn = row[0].rstrip(' ')
            name = fn + ' ' + sn
            # print(fn, sn, '*', name)
            if fn != 'forename':
                regstuds[name] = (regno, fn, sn)
    return regstuds
        
regstuds = importRegsFromCoTutor(regfile)

with open(logfile, newline='') as csvfile:
    reader = csv.reader(csvfile, delimiter=',', quotechar='"')
    for row in reader:
        when = row[0]
        who  = row[1]
        what = row[3]
        what = re.sub(' \(.*\)', '', what)
        if isInteresting(what):
            if isOther(what):
                if what not in otherInteresting:
                    otherInteresting = otherInteresting + [what]
            if who not in students and (who in regstuds or who == 'D D Freydenberger'):
                students[who] = True
        #TODO: refactor to use default dictionaries
            if who not in firstAccFor:
                firstAccFor[who] = dict()
            whats = [what]
            if isSheet(what) or isSpoiler(what):
                whats += [generalizedSheet(what)]
            for w in whats:
                if w not in firstAccFor[who]:
                    firstAccFor[who][w] = datetime.strptime(when, dateformat)
                else:
                    time = firstAccFor[who][w]
                    newTime = datetime.strptime(when, dateformat)
                    if newTime < time:
                        time = newTime
                    firstAccFor[who][w] = time
                    

gsd = dict()
spd = dict()
shd = dict()
for i in range(1,11):
    shd[i] = 'File: Sheet ' + str(i)
    spd[i] = shd[i] + ' with spoilers'
    gsd[i] = spd[i] + ' (or not)'

wb = Workbook()
desc = [gsd, shd, spd]
title = ['Any', 'NoSpoiler', 'Spoiler']
ws = [0] * (2*len(desc)+2)

ws[0] =  wb.active
ws[0].title = title[0]
ws[1] = wb.create_sheet(title=title[1])
ws[2] = wb.create_sheet(title=title[2])
for i in range(0,len(desc)):
    ws[i+len(desc)] = wb.create_sheet(title=title[i]+'NormDiff')
ws[2*len(desc)] = wb.create_sheet(title='Others')
ws[2*len(desc)+1] = wb.create_sheet(title='NormOthers')

outputDateFormat = "%d.%m.%Y"
date_style = NamedStyle(name='datetime', number_format='DD/MM/YY')
otherInteresting.sort()

for  j in range(0, 2*len(desc)):
    row = ['First name', 'Last name', 'regno']
    for i in range(1, numSheets + 1):
        row = row + [i]
    ws[j].append(row)
    for col in ws[j].iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)

for s in regstuds:
    for j in range(0, len(desc)):
        (regno, fn, sn) = regstuds[s]
        row = [fn, sn, regno]
        dayrow = row
        diffrow = row
        for i in range(1, numSheets + 1):
            dayri = ''
            if s in firstAccFor and desc[j][i] in firstAccFor[s]:
                dayri = firstAccFor[s][desc[j][i]].date()
                diffri = (examDate - firstAccFor[s][desc[j][i]].date()) / (examDate - firstAccFor[lecturer][desc[j][i]].date())
            else:
                diffri = 0
            dayrow = dayrow + [dayri]
            diffrow = diffrow + [diffri]
        ws[j].append(dayrow)
        ws[j+len(desc)].append(diffrow)

for j in range(0, len(desc)):
    for col in ws[j].iter_cols(min_col=4, min_row=2):
        for cell in col:
            cell.style = date_style
    for col in ws[j+len(desc)].iter_cols(min_col=4, min_row=2):
        for cell in col:
            cell.number_format = '0.00'

# make others, very copypasty
for j in range(2*len(desc), 2*len(desc)+2):
    row = ['First name', 'Last name', 'regno']
    for o in otherInteresting:
        lab = o.replace('File: ', '', 1).replace('Quiz: ', '', 1).replace('Folder: ', '', 1)
        row = row + [lab]
    ws[j].append(row)
    for col in ws[j].iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)

for s in regstuds:
     #ugly, but I am too lazy to deal with indentation right now.
    for j in range(2*len(desc), 2*len(desc)+1):
        (regno, fn, sn) = regstuds[s]
        row = [fn, sn, regno]
        dayrow = row
        diffrow = row
        for o in otherInteresting:
            dayri = ''
            if s in firstAccFor and o in firstAccFor[s]:
                dayri = firstAccFor[s][o].date()
                diffri = (examDate - firstAccFor[s][o].date()) / (examDate - firstAccFor[lecturer][o].date())
            else:
                diffri = 0
            dayrow = dayrow + [dayri]
            diffrow = diffrow + [diffri]
        ws[j].append(dayrow)
        ws[j+1].append(diffrow)
    for col in ws[j].iter_cols(min_col=4, min_row=2):
        for cell in col:
            cell.style = date_style
    for col in ws[j+1].iter_cols(min_col=4, min_row=2):
        for cell in col:
            cell.number_format = '0.00'

for i in range(0,len(ws)):
    ws[i].freeze_panes = "A2"

wb.save('out.xlsx')
